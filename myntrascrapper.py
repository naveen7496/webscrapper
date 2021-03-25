from bs4 import BeautifulSoup
import requests
import lxml
from openpyxl import load_workbook

html = requests.get('https://www.flipkart.com/clothing-and-accessories/topwear/pr?sid=clo,ash&p[]=facets.ideal_for%255B%255D%3DMen&p[]=facets.ideal_for%255B%255D%3Dmen&otracker=categorytree&fm=neo%2Fmerchandising&iid=M_c8c79dda-c902-4e1b-8607-c627f81ee3fe_2_372UD5BXDFYS_MC.6XNZG1FYFBZT&otracker=hp_rich_navigation_1_2.navigationCard.RICH_NAVIGATION_Fashion~Men%2527s%2BTop%2BWear_6XNZG1FYFBZT&cid=6XNZG1FYFBZT').text
soup = BeautifulSoup(html, 'lxml')
desc1 = soup.find_all('div', class_='_2WkVRV')
description = soup.find_all('a', class_='IRpwTa')
prices = soup.find_all('div', class_='_30jeq3')
for item in prices:
    price = item.text
    p = price[1:]
    print(p)
print(len(desc1))
print(len(description))
wb = load_workbook('Book1.xlsx')
sheet1 = wb.active
item_number = 0
row = 2
while item_number < len(description):   

    add_desc1 = desc1[item_number]
    add_description = description[item_number]
    add_price = prices[item_number]

    
    sheet1.cell(row, 1).value = add_desc1.text
    sheet1.cell(row, 2).value = add_description.text
    sheet1.cell(row, 3).value = add_price.text
    item_number +=1
    row += 1

wb.save('Book1.xlsx')

# a1 = sheet1.cell(1,1).value
# print(a1)
